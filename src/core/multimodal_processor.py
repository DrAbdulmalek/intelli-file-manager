"""معالج متعدد الوسائط - صور وصوت وفيديو"""
import os
import logging
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)


class MultimodalProcessor:
    """معالج ملفات متعددة الوسائط"""

    def __init__(self, model: str = None):
        self._model = model  # None = سيستخدم Config.ai_model
        self._ocr_available = False
        self._check_ocr()

    def _check_ocr(self):
        """التحقق من تثبيت OCR"""
        try:
            import pytesseract
            self._ocr_available = True
        except ImportError:
            logger.info("Tesseract OCR غير مثبت")

    def process_image(self, filepath: str) -> dict:
        """تحليل صورة ووصف محتواها"""
        path = Path(filepath)
        result = {"path": str(path), "type": "image"}

        try:
            from PIL import Image
            img = Image.open(filepath)
            result["width"] = img.width
            result["height"] = img.height
            result["format"] = img.format
            result["mode"] = img.mode
        except Exception as e:
            result["error"] = str(e)

        # OCR
        if self._ocr_available:
            try:
                import pytesseract
                text = pytesseract.image_to_string(filepath, lang="ara+eng")
                result["extracted_text"] = text.strip()
                result["ocr_success"] = True
            except Exception as e:
                result["ocr_success"] = False
                result["ocr_error"] = str(e)

        # وصف AI
        try:
            import ollama
            from ..core.config import Config
            config = Config()
            model_name = self._model or config.ai_model
            client = ollama.Client(host=config.ollama_url)
            # لنسخ الصورة إلى base64
            import base64
            with open(filepath, "rb") as f:
                img_b64 = base64.b64encode(f.read()).decode()
            response = client.chat(model_name, messages=[
                {"role": "user", "content": f"صِف هذه الصورة باختصار بالعربية: ",
                 "images": [img_b64]}
            ])
            result["ai_description"] = response["message"]["content"]
        except Exception:
            pass

        return result

    def process_audio(self, filepath: str) -> dict:
        """تحليل ملف صوتي واستخراج النص"""
        path = Path(filepath)
        result = {"path": str(path), "type": "audio"}

        try:
            import speech_recognition as sr
            recognizer = sr.Recognizer()
            with sr.AudioFile(filepath) as source:
                audio_data = recognizer.record(source, duration=120)
            text = recognizer.recognize_google(audio_data, language="ar-SA")
            result["transcript"] = text
            result["success"] = True
        except ImportError:
            result["error"] = "SpeechRecognition غير مثبت"
        except Exception as e:
            result["error"] = str(e)
            result["success"] = False

        return result

    def process_video(self, filepath: str) -> dict:
        """تحليل فيديو واستخراج معلومات"""
        path = Path(filepath)
        result = {"path": str(path), "type": "video", "name": path.name}

        try:
            # محاولة الحصول على معلومات باستخدام ffprobe
            import subprocess
            cmd = [
                "ffprobe", "-v", "quiet", "-print_format", "json",
                "-show_format", "-show_streams", str(filepath)
            ]
            proc = subprocess.run(cmd, capture_output=True, text=True, timeout=10)
            if proc.returncode == 0:
                import json
                info = json.loads(proc.stdout)
                if "format" in info:
                    fmt = info["format"]
                    result["duration"] = fmt.get("duration", "0")
                    result["bit_rate"] = fmt.get("bit_rate", "0")
                    result["format_name"] = fmt.get("format_name", "")
                    result["size"] = fmt.get("size", "0")
                if "streams" in info:
                    for stream in info["streams"]:
                        if stream.get("codec_type") == "video":
                            result["video_codec"] = stream.get("codec_name", "")
                            result["resolution"] = f"{stream.get('width')}x{stream.get('height')}"
                        elif stream.get("codec_type") == "audio":
                            result["audio_codec"] = stream.get("codec_name", "")
        except Exception as e:
            result["probe_error"] = str(e)

        return result

    def extract_text_from_pdf(self, filepath: str) -> str:
        """استخراج نص من PDF"""
        try:
            import pdfplumber
            with pdfplumber.open(filepath) as pdf:
                return "\n".join(p.extract_text() or "" for p in pdf.pages)
        except Exception as e:
            logger.error(f"خطأ في استخراج نص PDF: {e}")
            return ""

    def extract_text_from_docx(self, filepath: str) -> str:
        """استخراج نص من DOCX"""
        try:
            from docx import Document
            doc = Document(filepath)
            return "\n".join(p.text for p in doc.paragraphs)
        except Exception as e:
            logger.error(f"خطأ في استخراج نص DOCX: {e}")
            return ""

    def extract_text_from_xlsx(self, filepath: str) -> str:
        """استخراج نص من XLSX"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath)
            rows = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_text = " | ".join(str(c) if c else "" for c in row)
                    if row_text.strip():
                        rows.append(row_text)
            return "\n".join(rows[:500])
        except Exception as e:
            logger.error(f"خطأ في استخراج نص XLSX: {e}")
            return ""
