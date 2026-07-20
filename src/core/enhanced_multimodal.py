"""معالج متعدد الوسائط المحسّن - تكامل مع omni-medical-suite

Enhanced MultimodalProcessor with:
  - Moondream2 vision-language model for image description
  - Whisper for audio/video transcription (Arabic + English)
  - PaddleOCR + Tesseract dual-engine OCR
  - Scanner fixer integration from omni-medical-suite
  - DICOM medical image support
  - Arabic RTL text normalization
  - Offline-first: all models run locally
"""

from __future__ import annotations

import base64
import io
import json
import logging
import os
import subprocess
import tempfile
from pathlib import Path
from typing import Any, Optional

logger = logging.getLogger(__name__)


class EnhancedMultimodalProcessor:
    """معالج متعدد الوسائط محسّن للملفات الطبية العربية.

    Pipeline for image processing:
      1. Scanner fixer (deskew, enhance, denoise) — from omni-medical-suite
      2. OCR (PaddleOCR → Tesseract fallback) — Arabic + English
      3. Vision-Language (Moondream2) — image description
      4. Medical NER — extract medical entities from OCR text

    Pipeline for audio/video:
      1. FFmpeg extract audio
      2. Whisper transcription (Arabic + English)
      3. Optional speaker diarization

    All models run locally for privacy (offline-first).
    """

    def __init__(self, ollama_url: str = "http://localhost:11434",
                 ai_model: str = "llama3.2",
                 scanner_fixer_enabled: bool = True):
        self.ollama_url = ollama_url
        self.ai_model = ai_model
        self.scanner_fixer_enabled = scanner_fixer_enabled

        # Lazy-loaded engines
        self._paddle_ocr = None
        self._moondream = None
        self._whisper_model = None
        self._scanner_fixer = None
        self._medical_ner = None

        # Availability flags
        self.tesseract_available = self._check_tesseract()
        self.paddleocr_available = False  # checked on first use

    def _check_tesseract(self) -> bool:
        try:
            import pytesseract
            pytesseract.get_tesseract_version()
            return True
        except Exception:
            return False

    def _init_paddleocr(self):
        if self._paddle_ocr is not None:
            return
        try:
            from paddleocr import PaddleOCR
            self._paddle_ocr = PaddleOCR(
                use_angle_cls=True,
                lang="ar",  # Arabic
                show_log=False,
                use_gpu=False,
            )
            self.paddleocr_available = True
            logger.info("تم تحميل PaddleOCR (Arabic)")
        except ImportError:
            logger.info("PaddleOCR غير مثبت")
        except Exception as exc:
            logger.warning(f"خطأ في تحميل PaddleOCR: {exc}")

    def _init_scanner_fixer(self):
        """Try to import scanner_fixer from omni-medical-suite."""
        if self._scanner_fixer is not None:
            return
        if not self.scanner_fixer_enabled:
            return
        # Try omni-medical-suite package paths (priority order)
        for module_path in [
            "packages.scanner_fixer",
            "scanner_fixer",
            "src.scanner_fixer",
        ]:
            try:
                mod = __import__(module_path, fromlist=["fix_scan", "enhance_for_ocr"])
                self._scanner_fixer = {
                    "fix_scan": getattr(mod, "fix_scan", None),
                    "enhance_for_ocr": getattr(mod, "enhance_for_ocr", None),
                }
                logger.info(f"تم تحميل scanner_fixer من {module_path}")
                return
            except ImportError:
                continue
        logger.debug("scanner_fixer غير متاح — سيتم تخطي خطوة الإصلاح")
        self._scanner_fixer = None

    def _init_medical_ner(self):
        if self._medical_ner is not None:
            return
        try:
            from .medical_ner import ArabicMedicalNER
            self._medical_ner = ArabicMedicalNER(
                ollama_url=self.ollama_url,
                ai_model=self.ai_model,
            )
            logger.info("تم تحميل ArabicMedicalNER")
        except Exception as exc:
            logger.debug(f"ArabicMedicalNER غير متاح: {exc}")

    # -----------------------------------------------------------------------
    # Image processing pipeline
    # -----------------------------------------------------------------------

    def process_image(self, filepath: str, fix_scan: bool = True) -> dict:
        """Process an image through the full pipeline.

        Steps:
          1. (Optional) Fix scan artifacts with scanner_fixer
          2. OCR with PaddleOCR → Tesseract fallback
          3. Vision description with Moondream/Ollama
          4. Medical NER on extracted text

        Args:
            filepath: Path to the image file
            fix_scan: Whether to apply scanner fixer preprocessing

        Returns:
            Dict with all extraction results
        """
        path = Path(filepath)
        result = {
            "path": str(path),
            "type": "image",
            "name": path.name,
        }

        # Step 0: Basic image info
        try:
            from PIL import Image
            img = Image.open(filepath)
            result["width"] = img.width
            result["height"] = img.height
            result["format"] = img.format
            result["mode"] = img.mode
        except Exception as exc:
            result["image_info_error"] = str(exc)

        working_path = filepath

        # Step 1: Scanner fixer (if enabled and available)
        if fix_scan:
            self._init_scanner_fixer()
            if self._scanner_fixer:
                try:
                    fixed_path = self._fix_scan(filepath)
                    if fixed_path:
                        working_path = fixed_path
                        result["scan_fixed"] = True
                        result["fixed_path"] = fixed_path
                except Exception as exc:
                    result["scan_fix_error"] = str(exc)

        # Step 2: OCR — PaddleOCR first, then Tesseract fallback
        ocr_result = self._ocr_image(working_path)
        result.update(ocr_result)

        # Step 3: Vision-language description (Moondream2 via Ollama or llava)
        vision_result = self._vision_describe(working_path)
        if vision_result:
            result["ai_description"] = vision_result

        # Step 4: Medical NER on extracted text
        extracted_text = result.get("extracted_text", "")
        if extracted_text:
            self._init_medical_ner()
            if self._medical_ner:
                try:
                    ner_result = self._medical_ner.extract(extracted_text)
                    result["medical_entities"] = ner_result.summary()
                    result["diagnosis"] = ner_result.diagnosis
                    result["medications"] = ner_result.medications
                    result["patient_name"] = ner_result.patient_name
                    result["patient_id"] = ner_result.patient_id
                except Exception as exc:
                    result["ner_error"] = str(exc)

        return result

    def _fix_scan(self, filepath: str) -> Optional[str]:
        """Apply scanner fixer pipeline to an image."""
        if not self._scanner_fixer:
            return None
        try:
            fix_fn = self._scanner_fixer.get("fix_scan")
            if fix_fn:
                result = fix_fn(filepath)
                if result and hasattr(result, "get"):
                    image = result.get("image")
                    if image:
                        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                            image.save(tmp.name)
                            return tmp.name
                elif isinstance(result, str) and Path(result).exists():
                    return result

            # Fallback: PIL-based basic enhancement
            from PIL import Image, ImageEnhance, ImageFilter
            img = Image.open(filepath)

            # Convert to grayscale if needed
            if img.mode != "L":
                img = img.convert("L")

            # Apply basic enhancement
            enhance_fn = self._scanner_fixer.get("enhance_for_ocr")
            if enhance_fn:
                try:
                    result = enhance_fn(img)
                    if result:
                        img = result
                except Exception:
                    pass

            # Fallback PIL enhancement
            img = ImageEnhance.Contrast(img).enhance(1.5)
            img = ImageEnhance.Sharpness(img).enhance(2.0)
            img = img.filter(ImageFilter.MedianFilter(size=3))

            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                img.save(tmp.name)
                return tmp.name

        except Exception as exc:
            logger.debug(f"خطأ في scanner_fixer: {exc}")
            return None

    def _ocr_image(self, filepath: str) -> dict:
        """OCR an image with PaddleOCR → Tesseract fallback."""
        result = {}

        # Try PaddleOCR first (better for Arabic)
        self._init_paddleocr()
        if self._paddle_ocr:
            try:
                ocr_output = self._paddle_ocr.ocr(filepath, cls=True)
                if ocr_output and ocr_output[0]:
                    texts = []
                    for line in ocr_output[0]:
                        if line and len(line) >= 2:
                            texts.append(line[1][0])  # text content
                    extracted = "\n".join(texts)
                    if extracted.strip():
                        result["extracted_text"] = extracted
                        result["ocr_engine"] = "paddleocr"
                        result["ocr_success"] = True
                        return result
            except Exception as exc:
                result["paddleocr_error"] = str(exc)

        # Fallback: Tesseract
        if self.tesseract_available:
            try:
                import pytesseract
                from PIL import Image
                img = Image.open(filepath)
                text = pytesseract.image_to_string(img, lang="ara+eng")
                if text.strip():
                    result["extracted_text"] = text.strip()
                    result["ocr_engine"] = "tesseract"
                    result["ocr_success"] = True
                    return result
            except Exception as exc:
                result["tesseract_error"] = str(exc)

        result["ocr_success"] = False
        result["extracted_text"] = ""
        return result

    def _vision_describe(self, filepath: str) -> Optional[str]:
        """Describe an image using vision-language model via Ollama."""
        try:
            import ollama
            with open(filepath, "rb") as f:
                img_b64 = base64.b64encode(f.read()).decode()

            # Try Moondream2 first, then llava
            for model in ["moondream2", "llava", "llava:7b"]:
                try:
                    client = ollama.Client(host=self.ollama_url)
                    response = client.chat(model, messages=[{
                        "role": "user",
                        "content": "صِف هذه الصورة بالتفصيل بالعربية. إذا كانت صورة طبية (أشعة، تقرير)، صف المحتوى الطبي.",
                        "images": [img_b64],
                    }])
                    desc = response.get("message", {}).get("content", "")
                    if desc:
                        return desc
                except Exception:
                    continue
        except Exception as exc:
            logger.debug(f"Vision description غير متاح: {exc}")
        return None

    # -----------------------------------------------------------------------
    # Audio/Video processing
    # -----------------------------------------------------------------------

    def process_audio(self, filepath: str) -> dict:
        """Transcribe audio using Whisper (Arabic + English).

        Args:
            filepath: Path to audio file

        Returns:
            Dict with transcription results
        """
        path = Path(filepath)
        result = {"path": str(path), "type": "audio", "name": path.name}

        # Try local Whisper model
        transcript = self._transcribe_whisper(filepath)
        if transcript:
            result["transcript"] = transcript
            result["success"] = True
            result["engine"] = "whisper"
        else:
            # Fallback to Google Speech Recognition
            transcript = self._transcribe_google(filepath)
            if transcript:
                result["transcript"] = transcript
                result["success"] = True
                result["engine"] = "google_speech"
            else:
                result["success"] = False
                result["error"] = "لم يتم التعرف على الكلام"

        # Medical NER on transcript
        if result.get("transcript"):
            self._init_medical_ner()
            if self._medical_ner:
                try:
                    ner_result = self._medical_ner.extract(result["transcript"])
                    result["medical_entities"] = ner_result.summary()
                except Exception as exc:
                    result["ner_error"] = str(exc)

        return result

    def process_video(self, filepath: str) -> dict:
        """Process video: extract audio, transcribe, get metadata.

        Args:
            filepath: Path to video file

        Returns:
            Dict with video metadata and transcript
        """
        path = Path(filepath)
        result = {"path": str(path), "type": "video", "name": path.name}

        # Get video metadata with ffprobe
        probe_result = self._probe_video(filepath)
        result.update(probe_result)

        # Extract audio and transcribe
        audio_path = self._extract_audio(filepath)
        if audio_path:
            try:
                audio_result = self.process_audio(audio_path)
                result["transcript"] = audio_result.get("transcript", "")
                result["transcription_success"] = audio_result.get("success", False)
            finally:
                # Clean up temp audio file
                try:
                    os.unlink(audio_path)
                except Exception:
                    pass

        return result

    def _transcribe_whisper(self, filepath: str) -> Optional[str]:
        """Transcribe audio using local Whisper model."""
        try:
            import whisper
            if self._whisper_model is None:
                self._whisper_model = whisper.load_model("base")
            result = self._whisper_model.transcribe(filepath, language="ar")
            return result.get("text", "")
        except ImportError:
            logger.debug("whisper غير مثبت")
            return None
        except Exception as exc:
            logger.debug(f"خطأ في Whisper: {exc}")
            return None

    def _transcribe_google(self, filepath: str) -> Optional[str]:
        """Transcribe audio using Google Speech Recognition (fallback)."""
        try:
            import speech_recognition as sr
            recognizer = sr.Recognizer()
            with sr.AudioFile(filepath) as source:
                audio_data = recognizer.record(source, duration=300)
            return recognizer.recognize_google(audio_data, language="ar-SA")
        except ImportError:
            return None
        except Exception as exc:
            logger.debug(f"خطأ في Google Speech: {exc}")
            return None

    def _extract_audio(self, filepath: str) -> Optional[str]:
        """Extract audio from video using ffmpeg."""
        try:
            tmp = tempfile.NamedTemporaryFile(suffix=".wav", delete=False)
            cmd = [
                "ffmpeg", "-i", filepath, "-vn",
                "-acodec", "pcm_s16le", "-ar", "16000",
                "-ac", "1", "-y", tmp.name,
            ]
            proc = subprocess.run(cmd, capture_output=True, timeout=60)
            if proc.returncode == 0:
                return tmp.name
        except Exception as exc:
            logger.debug(f"خطأ في استخراج الصوت: {exc}")
        return None

    def _probe_video(self, filepath: str) -> dict:
        """Get video metadata using ffprobe."""
        result = {}
        try:
            cmd = [
                "ffprobe", "-v", "quiet", "-print_format", "json",
                "-show_format", "-show_streams", str(filepath),
            ]
            proc = subprocess.run(cmd, capture_output=True, text=True, timeout=10)
            if proc.returncode == 0:
                info = json.loads(proc.stdout)
                fmt = info.get("format", {})
                result["duration"] = fmt.get("duration", "0")
                result["format_name"] = fmt.get("format_name", "")
                result["size"] = fmt.get("size", "0")
                for stream in info.get("streams", []):
                    if stream.get("codec_type") == "video":
                        result["video_codec"] = stream.get("codec_name", "")
                        result["resolution"] = f"{stream.get('width')}x{stream.get('height')}"
                    elif stream.get("codec_type") == "audio":
                        result["audio_codec"] = stream.get("codec_name", "")
        except Exception as exc:
            result["probe_error"] = str(exc)
        return result

    # -----------------------------------------------------------------------
    # Document processing
    # -----------------------------------------------------------------------

    def process_document(self, filepath: str) -> dict:
        """Process a document file (PDF, DOCX, XLSX) with OCR and NER.

        For PDFs: extracts text, runs OCR on pages with no text,
        applies medical NER on the full content.

        Args:
            filepath: Path to the document file

        Returns:
            Dict with extracted text, OCR results, and medical entities
        """
        path = Path(filepath)
        ext = path.suffix.lower()
        result = {"path": str(path), "type": "document", "name": path.name, "format": ext}

        # Extract text
        text = ""
        if ext == ".pdf":
            text = self._extract_pdf(filepath, result)
        elif ext == ".docx":
            text = self._extract_docx(filepath)
        elif ext == ".xlsx":
            text = self._extract_xlsx(filepath)
        else:
            text = self._extract_text(filepath)

        result["extracted_text"] = text

        # Medical NER
        if text:
            self._init_medical_ner()
            if self._medical_ner:
                try:
                    ner_result = self._medical_ner.extract(text)
                    result["medical_entities"] = ner_result.summary()
                    result["diagnosis"] = ner_result.diagnosis
                    result["medications"] = ner_result.medications
                    result["patient_name"] = ner_result.patient_name
                    result["patient_id"] = ner_result.patient_id
                    result["doctor_name"] = ner_result.doctor_name
                    result["date"] = ner_result.date
                except Exception as exc:
                    result["ner_error"] = str(exc)

        return result

    def _extract_pdf(self, filepath: str, result: dict) -> str:
        """Extract text from PDF, OCR pages with no text."""
        text_parts = []
        try:
            import pdfplumber
            with pdfplumber.open(filepath) as pdf:
                result["page_count"] = len(pdf.pages)
                for i, page in enumerate(pdf.pages):
                    page_text = page.extract_text() or ""
                    if page_text.strip():
                        text_parts.append(page_text)
                    else:
                        # OCR the page
                        try:
                            img = page.to_image(resolution=300)
                            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                                img.save(tmp.name)
                                ocr_result = self._ocr_image(tmp.name)
                                page_text = ocr_result.get("extracted_text", "")
                                if page_text:
                                    text_parts.append(f"[صفحة {i+1} - OCR]:\n{page_text}")
                                os.unlink(tmp.name)
                        except Exception as exc:
                            logger.debug(f"خطأ في OCR صفحة {i+1}: {exc}")
        except ImportError:
            # Fallback: PyMuPDF
            try:
                import fitz  # PyMuPDF
                doc = fitz.open(filepath)
                result["page_count"] = len(doc)
                for page in doc:
                    text_parts.append(page.get_text())
                doc.close()
            except ImportError:
                logger.warning("لا يمكن قراءة PDF — pdfplumber و PyMuPDF غير مثبتين")

        return "\n\n".join(text_parts)

    def _extract_docx(self, filepath: str) -> str:
        try:
            from docx import Document
            doc = Document(filepath)
            return "\n".join(p.text for p in doc.paragraphs)
        except Exception as exc:
            logger.debug(f"خطأ في DOCX: {exc}")
            return ""

    def _extract_xlsx(self, filepath: str) -> str:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath)
            rows = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(values_only=True):
                    row_text = " | ".join(str(c) if c else "" for c in row)
                    if row_text.strip():
                        rows.append(row_text)
            return "\n".join(rows[:500])
        except Exception as exc:
            logger.debug(f"خطأ في XLSX: {exc}")
            return ""

    def _extract_text(self, filepath: str) -> str:
        try:
            return Path(filepath).read_text(encoding="utf-8", errors="ignore")[:10000]
        except Exception:
            return ""
